from collections.abc import AsyncIterable, AsyncIterator, Iterator
from dataclasses import dataclass
from datetime import date, datetime
import os
from pydantic import ValidationError
import shutil
import json as _json
import base64
import sys
import tempfile
from typing import Annotated, Optional, List, Any
from urllib.parse import urlparse

import openpyxl
import typer
import pyarrow as pa
import pyarrow.csv as pa_csv
import pyarrow.parquet as pq
from rich import print
from rich.console import Console
from rich.table import Table
import xlsxwriter

from meshagent.api import RequiredTable
from meshagent.api.http import new_client_session
from meshagent.agents.agent import install_required_table

from meshagent.cli.common_options import OutputFormatOption, ProjectIdOption, RoomOption
from meshagent.cli import async_typer
from meshagent.cli.helper import (
    get_client,
    print_json_table,
    resolve_project_id,
    resolve_room,
)
from meshagent.api.helpers import websocket_room_url
from meshagent.api import RoomClient, WebSocketClientProtocol
from meshagent.api.room_server_client import (
    DatasetIndexConfig,
    DatasetOptimizeConfig,
    DatasetSqlStatement,
    SqlTableReference,
)
from meshagent.api.sql import ALLOWED_DATA_TYPES, SchemaParseError, parse_table_schema
from meshagent.api import RoomException  # or wherever you defined it

app = async_typer.AsyncTyper(help="Manage dataset tables in a room")
branch_app = async_typer.AsyncTyper(help="Manage dataset branches in a room namespace.")
app.add_typer(
    branch_app,
    name="branch",
    help="Manage dataset branches in a room namespace.",
)

COLUMN_DEFINITIONS_HELP = (
    "Comma-separated column definitions. Example: "
    '"names vector(20) null, tags list(text), meta struct(owner text, score float)". '
    f"CLI shorthand types: {', '.join(ALLOWED_DATA_TYPES)}. "
    "Vector syntax: vector(size[, element_type]). "
    "List syntax: list(element_type). "
    "Struct syntax: struct(field_name type[, ...])."
)

SQL_OUTPUT_FORMATS = {"json", "table", "arrow", "csv", "tsv", "parquet", "excel"}
DATASET_IMPORT_FORMATS = {"auto", "json", "arrow", "csv", "tsv", "parquet", "excel"}
DATASET_IMPORT_MODES = {"create", "replace", "merge"}
EXCEL_MAX_CELL_CHARS = 32767
DATASET_IMPORT_BATCH_SIZE = 8192
_ARROW_EXTENSION_NAME_METADATA_KEY = b"ARROW:extension:name"
_ARROW_JSON_EXTENSION_NAME = b"arrow.json"


@dataclass(frozen=True, slots=True)
class _DatasetImportSource:
    schema: pa.Schema | None
    batches: AsyncIterable[pa.Table | pa.RecordBatch]


# ---------------------------
# Helpers
# ---------------------------


def _parse_json_arg(json_str: Optional[str], *, name: str) -> Any:
    if json_str is None:
        return None
    try:
        return _json.loads(json_str)
    except Exception as e:
        raise typer.BadParameter(f"Invalid JSON for {name}: {e}")


async def _load_json_file(path: Optional[str], *, name: str) -> Any:
    """
    Load JSON from a local file path or an HTTP(S) URL.
    """
    if path is None:
        return None

    try:
        parsed = urlparse(path)

        # URL case
        if parsed.scheme in ("http", "https"):
            async with new_client_session() as session:
                async with session.get(path) as resp:
                    if resp.status >= 400:
                        body = await resp.text()
                        raise RuntimeError(f"HTTP {resp.status}: {body}")
                    charset = resp.charset or "utf-8"
                    data = (await resp.read()).decode(charset)
                return _json.loads(data)

        # Local file case
        with open(path, "r", encoding="utf-8") as f:
            return _json.load(f)

    except Exception as e:
        raise typer.BadParameter(f"Unable to read {name} from {path}: {e}")


def _ns(namespace: Optional[List[str]]) -> Optional[List[str]]:
    return namespace or None


def _build_sql_table_refs(
    *,
    table: Optional[List[str]],
    tables_obj: Any,
    namespace: Optional[list[str]],
    branch: Optional[str],
    version: Optional[int],
) -> list[SqlTableReference]:
    table_refs: list[SqlTableReference] = []

    if table is not None:
        for table_name in table:
            table_refs.append(
                SqlTableReference(
                    name=table_name,
                    namespace=namespace,
                    branch=branch,
                    version=version,
                )
            )

    if tables_obj is not None:
        for idx, entry in enumerate(tables_obj):
            if not isinstance(entry, dict):
                raise typer.BadParameter(
                    f"table reference at index {idx} must be a JSON object"
                )

            table_ref_payload = dict(entry)
            if namespace is not None and "namespace" not in table_ref_payload:
                table_ref_payload["namespace"] = namespace
            if branch is not None and "branch" not in table_ref_payload:
                table_ref_payload["branch"] = branch
            if version is not None and "version" not in table_ref_payload:
                table_ref_payload["version"] = version

            table_refs.append(SqlTableReference.model_validate(table_ref_payload))

    return table_refs


def _sql_params_table(params_obj: Any) -> pa.Table | None:
    if params_obj is None:
        return None
    return pa.table({key: [value] for key, value in params_obj.items()})


def _coerce_record_list(*, value: Any, name: str) -> list[dict[str, Any]]:
    if not isinstance(value, list):
        raise typer.BadParameter(f"{name} expects a JSON list of records")

    records = list[dict[str, Any]]()
    for index, record in enumerate(value):
        if not isinstance(record, dict):
            raise typer.BadParameter(
                f"{name} expects record {index} to be a JSON object"
            )
        records.append(dict(record))
    return records


async def _record_chunks(
    records: list[dict[str, Any]], *, rows_per_chunk: int = 128
) -> AsyncIterator[list[dict[str, Any]]]:
    if rows_per_chunk <= 0:
        raise ValueError("rows_per_chunk must be greater than zero")
    for index in range(0, len(records), rows_per_chunk):
        yield records[index : index + rows_per_chunk]


def _indent_block(text: str, prefix: str) -> str:
    return "\n".join(f"{prefix}{line}" for line in text.splitlines())


def _json_safe_sql_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bytes):
        try:
            return value.decode("utf-8")
        except UnicodeDecodeError:
            return {"base64": base64.b64encode(value).decode("ascii")}
    if isinstance(value, dict):
        return {key: _json_safe_sql_value(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_json_safe_sql_value(item) for item in value]
    return value


def _is_arrow_json_field(field: pa.Field) -> bool:
    metadata = field.metadata or {}
    return (
        metadata.get(_ARROW_EXTENSION_NAME_METADATA_KEY) == _ARROW_JSON_EXTENSION_NAME
    )


def _decode_arrow_json_value(value: Any) -> Any:
    if isinstance(value, str):
        try:
            return _json.loads(value)
        except _json.JSONDecodeError:
            return value
    if isinstance(value, bytes):
        try:
            return _json.loads(value.decode("utf-8"))
        except (UnicodeDecodeError, _json.JSONDecodeError):
            return value
    return value


def _json_safe_sql_row(
    row: dict[str, Any], *, schema: pa.Schema | None = None
) -> dict[str, Any]:
    fields_by_name = (
        {field.name: field for field in schema} if schema is not None else {}
    )
    normalized = dict[str, Any]()
    for key, value in row.items():
        field = fields_by_name.get(key)
        if field is not None and _is_arrow_json_field(field):
            value = _decode_arrow_json_value(value)
        normalized[key] = _json_safe_sql_value(value)
    return normalized


def _sql_display_records(table: pa.Table) -> list[dict[str, Any]]:
    return [_json_safe_sql_row(row, schema=table.schema) for row in table.to_pylist()]


async def _print_row_batches(
    *,
    batches: AsyncIterable[Any],
    pretty: bool,
) -> None:
    typer.echo("[", nl=False)
    first = True
    async for batch in batches:
        rows = batch.to_pylist() if isinstance(batch, pa.Table) else batch
        schema = batch.schema if isinstance(batch, pa.Table) else None
        for row in rows:
            if first:
                if pretty:
                    typer.echo("\n", nl=False)
            else:
                typer.echo(",\n" if pretty else ",", nl=False)

            row = _json_safe_sql_row(row, schema=schema)
            payload = _json.dumps(row, indent=2 if pretty else None)
            typer.echo(_indent_block(payload, "  ") if pretty else payload, nl=False)
            first = False

    if pretty and not first:
        typer.echo("\n", nl=False)
    typer.echo("]")


def _normalize_sql_output_format(value: str) -> str:
    normalized = value.lower()
    if normalized not in SQL_OUTPUT_FORMATS:
        expected = "|".join(sorted(SQL_OUTPUT_FORMATS))
        raise typer.BadParameter(f"--format must be one of: {expected}")
    return normalized


def _normalize_dataset_import_format(value: str) -> str:
    normalized = value.lower()
    if normalized not in DATASET_IMPORT_FORMATS:
        expected = "|".join(sorted(DATASET_IMPORT_FORMATS))
        raise typer.BadParameter(f"--format must be one of: {expected}")
    return normalized


def _normalize_dataset_import_mode(value: str) -> str:
    normalized = value.lower()
    if normalized not in DATASET_IMPORT_MODES:
        expected = "|".join(sorted(DATASET_IMPORT_MODES))
        raise typer.BadParameter(f"--mode must be one of: {expected}")
    return normalized


def _infer_dataset_import_format(path: str) -> str:
    lower_path = path.lower()
    if lower_path.endswith((".arrow", ".ipc", ".feather")):
        return "arrow"
    if lower_path.endswith(".csv"):
        return "csv"
    if lower_path.endswith(".tsv"):
        return "tsv"
    if lower_path.endswith(".parquet"):
        return "parquet"
    if lower_path.endswith((".xlsx", ".xlsm")):
        return "excel"
    if lower_path.endswith(".json"):
        return "json"
    raise typer.BadParameter(
        "Could not infer --format from file extension; pass --format explicitly"
    )


def _open_output(path: str | None, *, binary: bool):
    if path is None:
        return sys.stdout.buffer if binary else sys.stdout
    if binary:
        return open(path, "wb")
    return open(path, "w", encoding="utf-8")


def _print_records_table(
    records: list[dict[str, Any]], *, output_path: str | None = None
) -> None:
    if output_path is None:
        print_json_table(records)
        return
    if not records:
        raise SystemExit("No rows to print")

    table = Table(show_header=True, header_style="bold magenta")
    columns = list(records[0])
    for column in columns:
        table.add_column(column.title())
    for row in records:
        table.add_row(*(str(row.get(column, "")) for column in columns))

    with open(output_path, "w", encoding="utf-8") as handle:
        Console(file=handle).print(table)


def _stringify_nested_columns(table: pa.Table) -> pa.Table:
    arrays: list[pa.Array | pa.ChunkedArray] = []
    fields: list[pa.Field] = []
    for field in table.schema:
        column = table[field.name]
        if pa.types.is_nested(field.type):
            values = [
                None if value is None else _json.dumps(value, separators=(",", ":"))
                for value in column.to_pylist()
            ]
            arrays.append(pa.array(values, type=pa.string()))
            fields.append(pa.field(field.name, pa.string(), nullable=field.nullable))
        else:
            arrays.append(column)
            fields.append(field)
    return pa.table(arrays, schema=pa.schema(fields))


async def _async_arrow_batches(
    batches: Iterator[pa.Table | pa.RecordBatch],
) -> AsyncIterator[pa.Table | pa.RecordBatch]:
    for batch in batches:
        yield batch


def _arrow_import_source(path: str) -> _DatasetImportSource:
    try:
        reader = pa.ipc.open_file(path)
        schema = reader.schema

        def batches() -> Iterator[pa.RecordBatch]:
            batch_reader = pa.ipc.open_file(path)
            for index in range(batch_reader.num_record_batches):
                yield batch_reader.get_batch(index)

        return _DatasetImportSource(
            schema=schema, batches=_async_arrow_batches(batches())
        )
    except (pa.ArrowInvalid, OSError):
        reader = pa.ipc.open_stream(path)
        schema = reader.schema

        def batches() -> Iterator[pa.RecordBatch]:
            yield from pa.ipc.open_stream(path)

        return _DatasetImportSource(
            schema=schema, batches=_async_arrow_batches(batches())
        )


def _csv_import_source(path: str, *, delimiter: str) -> _DatasetImportSource:
    parse_options = pa_csv.ParseOptions(delimiter=delimiter)
    reader = pa_csv.open_csv(path, parse_options=parse_options)
    schema = reader.schema

    def batches() -> Iterator[pa.RecordBatch]:
        yield from pa_csv.open_csv(path, parse_options=parse_options)

    return _DatasetImportSource(schema=schema, batches=_async_arrow_batches(batches()))


def _parquet_import_source(path: str, *, batch_size: int) -> _DatasetImportSource:
    parquet_file = pq.ParquetFile(path)
    schema = parquet_file.schema_arrow

    def batches() -> Iterator[pa.RecordBatch]:
        yield from pq.ParquetFile(path).iter_batches(batch_size=batch_size)

    return _DatasetImportSource(schema=schema, batches=_async_arrow_batches(batches()))


def _iter_json_records(path: str) -> Iterator[dict[str, Any]]:
    decoder = _json.JSONDecoder()

    def parse_line(line: str) -> dict[str, Any] | None:
        if line.strip() == "":
            return None
        value = _json.loads(line)
        if not isinstance(value, dict):
            raise typer.BadParameter("JSON import rows must be objects")
        return value

    with open(path, encoding="utf-8") as handle:
        buffer = ""
        eof = False

        def read_more() -> None:
            nonlocal buffer, eof
            chunk = handle.read(1024 * 1024)
            if chunk == "":
                eof = True
            else:
                buffer += chunk

        while not eof and buffer.strip() == "":
            read_more()
        stripped = buffer.lstrip()
        if stripped.startswith("["):
            buffer = stripped[1:]
            while True:
                buffer = buffer.lstrip()
                while buffer == "" and not eof:
                    read_more()
                    buffer = buffer.lstrip()
                if buffer.startswith("]"):
                    return
                try:
                    value, index = decoder.raw_decode(buffer)
                except _json.JSONDecodeError:
                    if eof:
                        raise
                    read_more()
                    continue
                if not isinstance(value, dict):
                    raise typer.BadParameter("JSON import rows must be objects")
                yield value
                buffer = buffer[index:].lstrip()
                while buffer == "" and not eof:
                    read_more()
                    buffer = buffer.lstrip()
                if buffer.startswith(","):
                    buffer = buffer[1:]
                    continue
                if buffer.startswith("]"):
                    return
                if eof and buffer.strip() == "":
                    raise typer.BadParameter("JSON array import ended before ']'")
                raise typer.BadParameter("JSON array import expected ',' or ']'")
        else:
            pending = buffer
            while pending != "":
                line, separator, rest = pending.partition("\n")
                if separator == "":
                    continuation = handle.readline()
                    if continuation == "":
                        value = parse_line(line)
                        if value is not None:
                            yield value
                        return
                    pending = line + continuation
                    continue
                value = parse_line(line)
                if value is not None:
                    yield value
                pending = rest
            for line in handle:
                value = parse_line(line)
                if value is not None:
                    yield value


def _json_import_source(path: str, *, batch_size: int) -> _DatasetImportSource:
    def batches() -> Iterator[pa.Table]:
        records: list[dict[str, Any]] = []
        for record in _iter_json_records(path):
            records.append(record)
            if len(records) >= batch_size:
                yield pa.Table.from_pylist(records)
                records = []
        if records:
            yield pa.Table.from_pylist(records)

    return _DatasetImportSource(schema=None, batches=_async_arrow_batches(batches()))


async def _excel_import_batches(
    path: str,
    *,
    sheet: str | None = None,
    batch_size: int,
) -> AsyncIterator[pa.Table]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet is None:
            worksheet = workbook[workbook.sheetnames[0]]
        else:
            worksheet = workbook[sheet]
        rows = worksheet.iter_rows(values_only=True)
        try:
            header_row = next(rows)
        except StopIteration:
            return
        columns = [
            str(value).strip()
            if value is not None and str(value).strip() != ""
            else None
            for value in header_row
        ]
        if all(column is None for column in columns):
            return
        records: list[dict[str, Any]] = []
        for row in rows:
            record = dict[str, Any]()
            for column, value in zip(columns, row):
                if column is not None:
                    record[column] = value
            if any(value is not None for value in record.values()):
                records.append(record)
            if len(records) >= batch_size:
                yield pa.Table.from_pylist(records)
                records = []
        if records:
            yield pa.Table.from_pylist(records)
    finally:
        workbook.close()


def _import_source(
    *,
    path: str,
    import_format: str,
    sheet: str | None = None,
    batch_size: int = DATASET_IMPORT_BATCH_SIZE,
) -> _DatasetImportSource:
    import_format = _normalize_dataset_import_format(import_format)
    if import_format == "auto":
        import_format = _infer_dataset_import_format(path)
    if import_format == "arrow":
        return _arrow_import_source(path)
    if import_format == "csv":
        return _csv_import_source(path, delimiter=",")
    if import_format == "tsv":
        return _csv_import_source(path, delimiter="\t")
    if import_format == "parquet":
        return _parquet_import_source(path, batch_size=batch_size)
    if import_format == "json":
        return _json_import_source(path, batch_size=batch_size)
    if import_format == "excel":
        return _DatasetImportSource(
            schema=None,
            batches=_excel_import_batches(path, sheet=sheet, batch_size=batch_size),
        )
    raise AssertionError(f"unhandled dataset import format: {import_format}")


def _excel_cell_value(value: Any) -> Any:
    if isinstance(value, dict | list):
        value = _json.dumps(value, separators=(",", ":"))
    elif isinstance(value, bytes):
        value = value.hex()
    if isinstance(value, str) and len(value) > EXCEL_MAX_CELL_CHARS:
        return value[:EXCEL_MAX_CELL_CHARS]
    return value


def _create_excel_workbook(path: str) -> xlsxwriter.Workbook:
    return xlsxwriter.Workbook(path, {"constant_memory": True})


def _write_or_stream_excel_workbook(
    output_path: str | None, write_workbook: Any
) -> None:
    if output_path is not None:
        workbook = _create_excel_workbook(output_path)
        try:
            write_workbook(workbook)
        finally:
            workbook.close()
        return

    temporary = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    temporary_path = temporary.name
    temporary.close()
    try:
        workbook = _create_excel_workbook(temporary_path)
        try:
            write_workbook(workbook)
        finally:
            workbook.close()
        with open(temporary_path, "rb") as handle:
            shutil.copyfileobj(handle, sys.stdout.buffer)
    finally:
        os.unlink(temporary_path)


def _write_excel_rows(
    *,
    workbook: xlsxwriter.Workbook,
    worksheet_name: str,
    schema: pa.Schema,
    tables: list[pa.Table],
) -> None:
    worksheet = workbook.add_worksheet(worksheet_name)
    header_format = workbook.add_format({"bold": True})
    for column_index, field in enumerate(schema):
        worksheet.write(0, column_index, field.name, header_format)

    row_index = 1
    for table in tables:
        for row in table.to_pylist():
            for column_index, field in enumerate(schema):
                worksheet.write(
                    row_index, column_index, _excel_cell_value(row[field.name])
                )
            row_index += 1

    if len(schema) > 0:
        worksheet.autofilter(0, 0, max(row_index - 1, 0), len(schema) - 1)
        worksheet.freeze_panes(1, 0)


async def _write_sql_json_output(
    *,
    batches: AsyncIterable[Any],
    output_path: str | None,
    pretty: bool,
) -> None:
    if output_path is None:
        await _print_row_batches(batches=batches, pretty=pretty)
        return

    with open(output_path, "w", encoding="utf-8") as handle:
        handle.write("[")
        first = True
        async for batch in batches:
            rows = batch.to_pylist() if isinstance(batch, pa.Table) else batch
            schema = batch.schema if isinstance(batch, pa.Table) else None
            for row in rows:
                if first:
                    if pretty:
                        handle.write("\n")
                else:
                    handle.write(",\n" if pretty else ",")
                row = _json_safe_sql_row(row, schema=schema)
                payload = _json.dumps(row, indent=2 if pretty else None)
                handle.write(_indent_block(payload, "  ") if pretty else payload)
                first = False
        if pretty and not first:
            handle.write("\n")
        handle.write("]\n")


async def _write_sql_query_output(
    *,
    batches: AsyncIterable[pa.Table],
    schema: pa.Schema,
    output_format: str,
    output_path: str | None,
    pretty: bool,
) -> None:
    output_format = _normalize_sql_output_format(output_format)
    if output_format == "json":
        await _write_sql_json_output(
            batches=batches,
            output_path=output_path,
            pretty=pretty,
        )
        return
    if output_format == "table":
        tables = [batch async for batch in batches]
        table = (
            pa.concat_tables(tables) if tables else pa.Table.from_batches([], schema)
        )
        _print_records_table(_sql_display_records(table), output_path=output_path)
        return
    if output_format == "excel":
        tables = [batch async for batch in batches]

        def write_workbook(workbook: xlsxwriter.Workbook) -> None:
            _write_excel_rows(
                workbook=workbook,
                worksheet_name="Query",
                schema=schema,
                tables=tables,
            )

        _write_or_stream_excel_workbook(output_path, write_workbook)
        return

    binary = output_format in {"arrow", "csv", "tsv", "parquet"}
    sink = _open_output(output_path, binary=binary)
    close_sink = output_path is not None
    try:
        if output_format == "arrow":
            with pa.ipc.new_file(sink, schema) as writer:
                async for batch in batches:
                    writer.write_table(batch)
        elif output_format == "parquet":
            writer = pq.ParquetWriter(sink, schema)
            try:
                async for batch in batches:
                    writer.write_table(batch)
            finally:
                writer.close()
        elif output_format in {"csv", "tsv"}:
            write_options = pa_csv.WriteOptions(
                delimiter="\t" if output_format == "tsv" else ","
            )
            csv_schema = _stringify_nested_columns(
                pa.Table.from_batches([], schema=schema)
            ).schema
            with pa_csv.CSVWriter(
                sink, csv_schema, write_options=write_options
            ) as writer:
                async for batch in batches:
                    writer.write_table(_stringify_nested_columns(batch))
        else:
            raise AssertionError(f"unhandled SQL output format: {output_format}")
    finally:
        if close_sink:
            sink.close()


def _write_sql_statement_output(
    *,
    rows_affected: int,
    output_format: str,
    output_path: str | None,
    pretty: bool,
) -> None:
    output_format = _normalize_sql_output_format(output_format)
    table = pa.table({"rows_affected": [rows_affected]})
    if output_format == "json":
        payload = _json.dumps(
            {"rows_affected": rows_affected},
            indent=2 if pretty else None,
        )
        if output_path is None:
            typer.echo(payload)
        else:
            with open(output_path, "w", encoding="utf-8") as handle:
                handle.write(f"{payload}\n")
    elif output_format == "table":
        _print_records_table(table.to_pylist(), output_path=output_path)
    elif output_format == "excel":
        _write_or_stream_excel_workbook(
            output_path,
            lambda workbook: _write_excel_rows(
                workbook=workbook,
                worksheet_name="Statement",
                schema=table.schema,
                tables=[table],
            ),
        )
    elif output_format == "arrow":
        sink = _open_output(output_path, binary=True)
        try:
            with pa.ipc.new_file(sink, table.schema) as writer:
                writer.write_table(table)
        finally:
            if output_path is not None:
                sink.close()
    elif output_format == "parquet":
        sink = _open_output(output_path, binary=True)
        try:
            pq.write_table(table, sink)
        finally:
            if output_path is not None:
                sink.close()
    elif output_format in {"csv", "tsv"}:
        sink = _open_output(output_path, binary=True)
        try:
            pa_csv.write_csv(
                table,
                sink,
                write_options=pa_csv.WriteOptions(
                    delimiter="\t" if output_format == "tsv" else ","
                ),
            )
        finally:
            if output_path is not None:
                sink.close()


NamespaceOption = Annotated[
    Optional[List[str]],
    typer.Option(
        "--namespace",
        "-n",
        help="Namespace path segments (repeatable). Example: -n prod -n analytics",
    ),
]

BranchOption = Annotated[
    Optional[str],
    typer.Option("--branch", help="Dataset branch name (defaults to main)"),
]

VersionOption = Annotated[
    Optional[int],
    typer.Option(
        "--version",
        "-v",
        help="Historical table version to read (defaults to latest on the branch)",
    ),
]


# ---------------------------
# Commands
# ---------------------------


@app.async_command("tables", help="List dataset tables in a room.", hidden=True)
@app.async_command("table", help="List dataset tables in a room.")
async def list_tables(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            tables = await client.datasets.list_tables(
                namespace=_ns(namespace),
                branch=branch,
            )
            if not tables:
                print("[bold yellow]No tables found.[/bold yellow]")
            else:
                for t in tables:
                    print(t)

    except RoomException as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command("inspect", help="Inspect a table schema in a room dataset.")
async def inspect(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
    version: VersionOption = None,
    json: Annotated[
        bool, typer.Option("--json", help="Output raw schema JSON")
    ] = False,
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            schema = await client.datasets.inspect(
                table=table,
                namespace=_ns(namespace),
                branch=branch,
                version=version,
            )

            if json:
                print(schema.to_string())
            else:
                print(f"[bold]{table}[/bold]")
                print(schema.to_string())

    except RoomException as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command(
    "install", help="Install required tables from a requirements JSON file."
)
async def install_requirements(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    file: Annotated[
        Optional[str], typer.Option("--file", help="Path to requirements JSON file")
    ] = None,
):
    """
    Create a dataset from a json file containing a list of RequiredTables.
    """
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        requirements = await _load_json_file(file, name="--file")

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            for rt in requirements["tables"]:
                rt = RequiredTable.from_json(rt)
                print(f"installing table {rt.name} in namespace {rt.namespace}")
                await install_required_table(room=client, table=rt)

    except RoomException as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command(
    "create",
    help="Create a room dataset table with optional Arrow schema and seed data.",
)
async def create_table(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    mode: Annotated[
        str, typer.Option("--mode", help="create | overwrite | create_if_not_exists")
    ] = "create",
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
    columns: Annotated[
        Optional[str],
        typer.Option(
            "--columns",
            "-c",
            help=COLUMN_DEFINITIONS_HELP,
        ),
    ] = None,
    data_json: Annotated[
        Optional[str], typer.Option("--data-json", help="Initial rows (JSON list)")
    ] = None,
    data_file: Annotated[
        Optional[str],
        typer.Option("--data-file", help="Path to JSON file with initial rows"),
    ] = None,
):
    """
    Create a table with optional Arrow schema + optional initial data.

    Column definitions via --columns/-c use SQL-like syntax:
      names vector(20) null, tags list(text), meta struct(owner text, score float)

    Allowed types: int, bool, date, timestamp, float, text, json, binary, uuid, vector, list, struct.
    Vector syntax: vector(size[, element_type]).
    List syntax: list(element_type).
    Struct syntax: struct(field_name type[, ...]).
    """
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        data_obj = _parse_json_arg(data_json, name="--data-json")
        data_obj = (
            data_obj
            if data_obj is not None
            else await _load_json_file(data_file, name="--data-file")
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            arrow_schema: pa.Schema | None = None
            if columns is not None:
                try:
                    arrow_schema = parse_table_schema(columns)
                except SchemaParseError as e:
                    raise typer.BadParameter(str(e))

            if data_obj is not None:
                records = _coerce_record_list(value=data_obj, name="create")
                if arrow_schema is not None:
                    await client.datasets.create_table_with_schema(
                        name=table,
                        schema=arrow_schema,
                        data=pa.Table.from_pylist(records, schema=arrow_schema),
                        mode=mode,  # type: ignore
                        namespace=_ns(namespace),
                        branch=branch,
                    )
                else:
                    await client.datasets.create_table_from_json_data(
                        name=table,
                        data=records,
                        mode=mode,  # type: ignore
                        namespace=_ns(namespace),
                        branch=branch,
                    )
            elif arrow_schema is not None:
                await client.datasets.create_table_with_schema(
                    name=table,
                    schema=arrow_schema,
                    mode=mode,  # type: ignore
                    namespace=_ns(namespace),
                    branch=branch,
                )
            else:
                raise typer.BadParameter(
                    "Provide --columns for an Arrow schema or --data-json/--data-file for JSON data"
                )

            print(f"[bold green]Created table:[/bold green] {table}")

    except (RoomException, ValidationError) as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command(
    "import",
    help="Import a local Arrow, CSV, TSV, Parquet, JSON, or Excel file into a table.",
)
async def import_table(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    file: Annotated[
        str,
        typer.Option(
            ...,
            "--file",
            "-f",
            help="Local file to import",
        ),
    ],
    mode: Annotated[
        str,
        typer.Option("--mode", help="Import mode: create, replace, merge"),
    ] = "create",
    import_format: Annotated[
        str,
        typer.Option(
            "--format",
            help="Input format: auto, json, arrow, csv, tsv, parquet, excel",
        ),
    ] = "auto",
    on: Annotated[
        Optional[str],
        typer.Option("--on", help="Column to match when --mode merge"),
    ] = None,
    sheet: Annotated[
        Optional[str],
        typer.Option("--sheet", help="Excel worksheet name for --format excel"),
    ] = None,
    batch_size: Annotated[
        int,
        typer.Option(
            "--batch-size",
            help="Rows per imported batch for Parquet, JSON, and Excel",
        ),
    ] = DATASET_IMPORT_BATCH_SIZE,
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
):
    account_client = await get_client()
    try:
        mode = _normalize_dataset_import_mode(mode)
        import_format = _normalize_dataset_import_format(import_format)
        if batch_size <= 0:
            raise typer.BadParameter("--batch-size must be greater than zero")
        if mode == "merge":
            if on is None or on.strip() == "":
                raise typer.BadParameter("--mode merge requires --on")
        elif on is not None:
            raise typer.BadParameter("--on can only be used with --mode merge")

        source = _import_source(
            path=file,
            import_format=import_format,
            sheet=sheet,
            batch_size=batch_size,
        )

        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            if mode == "merge":
                await client.datasets.merge_stream(
                    table=table,
                    on=on.strip() if on is not None else "",
                    chunks=source.batches,
                    namespace=_ns(namespace),
                    branch=branch,
                )
                print(
                    f"[bold green]Merged import[/bold green] from {file} into {table} on {on}"
                )
            else:
                await client.datasets.create_table_from_data_stream(
                    name=table,
                    chunks=source.batches,
                    schema=source.schema,
                    mode="overwrite" if mode == "replace" else "create",
                    namespace=_ns(namespace),
                    branch=branch,
                )
                action = "Replaced" if mode == "replace" else "Created"
                print(f"[bold green]{action} table[/bold green] {table} from {file}")

    except (RoomException, typer.BadParameter, ValidationError, OSError, KeyError) as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command("drop", help="Drop a room dataset table.")
async def drop_table(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
    ignore_missing: Annotated[
        bool, typer.Option("--ignore-missing", help="Ignore missing table")
    ] = False,
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            await client.datasets.drop_table(
                name=table,
                ignore_missing=ignore_missing,
                namespace=_ns(namespace),
                branch=branch,
            )
            print(f"[bold green]Dropped table:[/bold green] {table}")

    except RoomException as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command("add-columns", help="Add columns to a room dataset table.")
async def add_columns(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
    columns: Annotated[
        Optional[str],
        typer.Option(
            "--columns",
            "-c",
            help=COLUMN_DEFINITIONS_HELP,
        ),
    ] = None,
    columns_json: Annotated[
        Optional[str],
        typer.Option(
            "--columns-json",
            help=(
                "JSON object of new columns mapped to SQL default expressions "
                '(e.g. \'{"created_at":"now()"}\').'
            ),
        ),
    ] = None,
):
    """
    Add columns. JSON supports server default SQL expression strings:
      {"col":"'default'"}

    Column definitions via --columns/-c use SQL-like syntax:
      names vector(20) null, tags list(text), meta struct(owner text, score float)

    Allowed types: int, bool, date, timestamp, float, text, json, binary, uuid, vector, list, struct.
    Vector syntax: vector(size[, element_type]).
    List syntax: list(element_type).
    Struct syntax: struct(field_name type[, ...]).
    """
    account_client = await get_client()
    try:
        if columns and columns_json:
            raise typer.BadParameter("Use --columns or --columns-json, not both")
        if columns is None and columns_json is None:
            raise typer.BadParameter("Provide --columns or --columns-json")

        cols_obj = _parse_json_arg(columns_json, name="--columns-json")
        if columns_json is not None and not isinstance(cols_obj, dict):
            raise typer.BadParameter("--columns-json must be a JSON object")

        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            if columns is not None:
                try:
                    new_cols = parse_table_schema(columns)
                except SchemaParseError as e:
                    raise typer.BadParameter(str(e))
            else:
                new_cols = {}
                for k, v in cols_obj.items():
                    if not isinstance(v, str):
                        raise typer.BadParameter(
                            "--columns-json values must be SQL expression strings"
                        )
                    new_cols[k] = v

            await client.datasets.add_columns(
                table=table,
                new_columns=new_cols,
                namespace=_ns(namespace),
                branch=branch,
            )
            print(f"[bold green]Added columns to[/bold green] {table}")

    except (RoomException, typer.BadParameter, ValidationError) as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command("drop-columns", help="Drop columns from a room dataset table.")
async def drop_columns(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
    columns: Annotated[
        List[str],
        typer.Option(..., "--column", "-c", help="Column to drop (repeatable)"),
    ] = None,
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            await client.datasets.drop_columns(
                table=table,
                columns=columns,
                namespace=_ns(namespace),
                branch=branch,
            )
            print(f"[bold green]Dropped columns from[/bold green] {table}")

    except RoomException as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command("insert", help="Insert records into a room dataset table.")
async def insert(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
    json: Annotated[
        Optional[str], typer.Option("--json", help="JSON list of records")
    ] = None,
    file: Annotated[
        Optional[str],
        typer.Option("--file", "-f", help="Path to JSON file (list of records)"),
    ] = None,
):
    account_client = await get_client()
    try:
        records = _parse_json_arg(json, name="--json")
        records = (
            records
            if records is not None
            else await _load_json_file(file, name="--file")
        )
        records = _coerce_record_list(value=records, name="insert")

        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            await client.datasets.insert_stream(
                table=table,
                chunks=[pa.Table.from_pylist(records)],
                namespace=_ns(namespace),
                branch=branch,
            )
            print(
                f"[bold green]Inserted[/bold green] {len(records)} record(s) into {table}"
            )

    except (RoomException, typer.BadParameter) as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command("merge", help="Upsert records into a room dataset table.")
async def merge(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    on: Annotated[str, typer.Option(..., "--on", help="Column to match for upsert")],
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
    json: Annotated[
        Optional[str], typer.Option("--json", help="JSON records (list)")
    ] = None,
    file: Annotated[
        Optional[str], typer.Option("--file", "-f", help="Path to JSON file (list)")
    ] = None,
):
    account_client = await get_client()
    try:
        records = _parse_json_arg(json, name="--json")
        records = (
            records
            if records is not None
            else await _load_json_file(file, name="--file")
        )
        records = _coerce_record_list(value=records, name="merge")

        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            await client.datasets.merge_stream(
                table=table,
                on=on,
                chunks=[pa.Table.from_pylist(records)],
                namespace=_ns(namespace),
                branch=branch,
            )
            print(
                f"[bold green]Merged[/bold green] {len(records)} record(s) into {table} on {on}"
            )

    except (RoomException, typer.BadParameter) as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command("update", help="Update rows in a room dataset table.")
async def update(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    where: Annotated[
        str, typer.Option(..., "--where", help='SQL WHERE clause, e.g. "id = 1"')
    ],
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
    values_json: Annotated[
        str,
        typer.Option(
            "--values-json",
            help='JSON object of update values; use {"column":{"expression":"..."}} for expressions',
        ),
    ],
):
    account_client = await get_client()
    try:
        values = _parse_json_arg(values_json, name="--values-json")

        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            await client.datasets.update(
                table=table,
                where=where,
                values=values,
                namespace=_ns(namespace),
                branch=branch,
            )
            print(f"[bold green]Updated[/bold green] {table} where {where}")

    except (RoomException, typer.BadParameter) as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command("delete", help="Delete rows from a room dataset table.")
async def delete(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    where: Annotated[str, typer.Option(..., "--where", help="SQL WHERE clause")],
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            await client.datasets.delete(
                table=table,
                where=where,
                namespace=_ns(namespace),
                branch=branch,
            )
            print(f"[bold green]Deleted[/bold green] from {table} where {where}")

    except RoomException as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command("search", help="Search rows in a room dataset table.")
async def search(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
    version: VersionOption = None,
    text: Annotated[
        Optional[str], typer.Option("--text", help="Full-text query")
    ] = None,
    vector_json: Annotated[
        Optional[str], typer.Option("--vector-json", help="Vector JSON array")
    ] = None,
    where: Annotated[
        Optional[str], typer.Option("--where", help="SQL WHERE clause")
    ] = None,
    where_json: Annotated[
        Optional[str],
        typer.Option("--where-json", help="JSON object converted to equality ANDs"),
    ] = None,
    select: Annotated[
        Optional[List[str]],
        typer.Option("--select", help="Columns to select (repeatable)"),
    ] = None,
    limit: Annotated[
        Optional[int], typer.Option("--limit", help="Max rows to return")
    ] = None,
    offset: Annotated[
        Optional[int], typer.Option("--offset", help="Rows to skip")
    ] = None,
    pretty: Annotated[
        bool, typer.Option("--pretty/--no-pretty", help="Pretty-print JSON")
    ] = True,
):
    account_client = await get_client()
    try:
        vec = _parse_json_arg(vector_json, name="--vector-json")
        if vec is not None and not isinstance(vec, list):
            raise typer.BadParameter("--vector-json must be a JSON array")
        wj = _parse_json_arg(where_json, name="--where-json")
        if wj is not None and not isinstance(wj, dict):
            raise typer.BadParameter("--where-json must be a JSON object")

        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            await _print_row_batches(
                batches=client.datasets.search_stream(
                    table=table,
                    text=text,
                    vector=vec,
                    where=(wj if wj is not None else where),
                    select=list(select) if select else None,
                    limit=limit,
                    offset=offset,
                    namespace=_ns(namespace),
                    branch=branch,
                    version=version,
                ),
                pretty=pretty,
            )

    except (RoomException, typer.BadParameter) as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command("sql", help="Execute SQL against room dataset tables.")
async def sql(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    query: Annotated[
        str,
        typer.Option(
            ...,
            "--query",
            "-q",
            help="SQL query to execute",
        ),
    ],
    table: Annotated[
        Optional[List[str]],
        typer.Option(
            "--table",
            "-t",
            help="Table name to register in SQL context (repeatable)",
        ),
    ] = None,
    namespace: NamespaceOption = None,
    tables_json: Annotated[
        Optional[str],
        typer.Option(
            "--tables-json",
            help=(
                "JSON array of table refs "
                '(e.g. \'[{"name":"users","alias":"u","namespace":["prod"]}]\')'
            ),
        ),
    ] = None,
    tables_file: Annotated[
        Optional[str],
        typer.Option(
            "--tables-file",
            help="Path/URL to JSON array of table refs (same format as --tables-json)",
        ),
    ] = None,
    params_json: Annotated[
        Optional[str],
        typer.Option(
            "--params-json",
            help="JSON object of SQL parameters for DataFusion param binding",
        ),
    ] = None,
    params_file: Annotated[
        Optional[str],
        typer.Option(
            "--params-file",
            help="Path/URL to JSON object of SQL parameters",
        ),
    ] = None,
    branch: BranchOption = None,
    version: VersionOption = None,
    pretty: Annotated[
        bool, typer.Option("--pretty/--no-pretty", help="Pretty-print JSON")
    ] = True,
    output_format: Annotated[
        str,
        typer.Option(
            "--format",
            help="Output format: table, json, arrow, csv, tsv, parquet, excel",
        ),
    ] = "table",
    output: Annotated[
        Optional[str],
        typer.Option(
            "--output",
            "-o",
            help="Write output to this file path instead of stdout",
        ),
    ] = None,
):
    """
    Execute SQL against one or more room dataset tables.

    You can pass table names with --table/-t, and optionally provide detailed table
    references with --tables-json/--tables-file for alias/namespace control.
    """
    account_client = await get_client()
    try:
        if query.strip() == "":
            raise typer.BadParameter("--query cannot be empty")
        output_format = _normalize_sql_output_format(output_format)

        if tables_json is not None and tables_file is not None:
            raise typer.BadParameter("Use --tables-json or --tables-file, not both")
        if params_json is not None and params_file is not None:
            raise typer.BadParameter("Use --params-json or --params-file, not both")

        tables_obj = _parse_json_arg(tables_json, name="--tables-json")
        tables_obj = (
            tables_obj
            if tables_obj is not None
            else await _load_json_file(tables_file, name="--tables-file")
        )
        if tables_obj is not None and not isinstance(tables_obj, list):
            raise typer.BadParameter(
                "--tables-json/--tables-file must be a JSON array of table references"
            )

        params_obj = _parse_json_arg(params_json, name="--params-json")
        params_obj = (
            params_obj
            if params_obj is not None
            else await _load_json_file(params_file, name="--params-file")
        )
        if params_obj is not None and not isinstance(params_obj, dict):
            raise typer.BadParameter(
                "--params-json/--params-file must be a JSON object"
            )

        resolved_namespace = _ns(namespace)
        table_refs = _build_sql_table_refs(
            table=table,
            tables_obj=tables_obj,
            namespace=resolved_namespace,
            branch=branch,
            version=version,
        )
        params_table = _sql_params_table(params_obj)

        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            result = await client.datasets.execute_sql(
                query=query,
                tables=table_refs or None,
                params=params_table,
                namespace=resolved_namespace,
                branch=branch,
            )
            if isinstance(result, DatasetSqlStatement):
                _write_sql_statement_output(
                    rows_affected=result.rows_affected,
                    output_format=output_format,
                    output_path=output,
                    pretty=pretty,
                )
            else:
                try:
                    await _write_sql_query_output(
                        batches=client.datasets.read_sql_query(
                            query_id=result.query_id,
                        ),
                        schema=result.schema,
                        output_format=output_format,
                        output_path=output,
                        pretty=pretty,
                    )
                finally:
                    await client.datasets.close_sql_query(query_id=result.query_id)

    except (RoomException, typer.BadParameter, ValidationError) as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command("optimize", help="Optimize a room dataset table.")
async def optimize(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    compact_files: Annotated[
        Optional[bool],
        typer.Option("--compact-files/--no-compact-files"),
    ] = None,
    optimize_indices: Annotated[
        Optional[bool],
        typer.Option("--optimize-indices/--no-optimize-indices"),
    ] = None,
    cleanup_old_versions: Annotated[
        Optional[bool],
        typer.Option("--cleanup-old-versions/--no-cleanup-old-versions"),
    ] = None,
    target_rows_per_fragment: Annotated[
        Optional[int], typer.Option("--target-rows-per-fragment")
    ] = None,
    max_rows_per_group: Annotated[
        Optional[int], typer.Option("--max-rows-per-group")
    ] = None,
    max_bytes_per_file: Annotated[
        Optional[int], typer.Option("--max-bytes-per-file")
    ] = None,
    materialize_deletions: Annotated[
        Optional[bool],
        typer.Option("--materialize-deletions/--no-materialize-deletions"),
    ] = None,
    materialize_deletions_threshold: Annotated[
        Optional[float], typer.Option("--materialize-deletions-threshold")
    ] = None,
    defer_index_remap: Annotated[
        Optional[bool],
        typer.Option("--defer-index-remap/--no-defer-index-remap"),
    ] = None,
    num_threads: Annotated[Optional[int], typer.Option("--num-threads")] = None,
    batch_size: Annotated[Optional[int], typer.Option("--batch-size")] = None,
    compaction_mode: Annotated[
        Optional[str],
        typer.Option(
            "--compaction-mode",
            help="Compaction mode: reencode, try_binary_copy, or force_binary_copy",
        ),
    ] = None,
    binary_copy_read_batch_bytes: Annotated[
        Optional[int], typer.Option("--binary-copy-read-batch-bytes")
    ] = None,
    num_indices_to_merge: Annotated[
        Optional[int], typer.Option("--num-indices-to-merge")
    ] = None,
    index_names: Annotated[
        Optional[List[str]],
        typer.Option("--index-name", help="Index name to optimize. Repeatable."),
    ] = None,
    retrain: Annotated[Optional[bool], typer.Option("--retrain/--no-retrain")] = None,
    older_than_seconds: Annotated[
        Optional[float], typer.Option("--older-than-seconds")
    ] = None,
    retain_versions: Annotated[Optional[int], typer.Option("--retain-versions")] = None,
    delete_unverified: Annotated[
        Optional[bool], typer.Option("--delete-unverified/--keep-unverified")
    ] = None,
    error_if_tagged_old_versions: Annotated[
        Optional[bool],
        typer.Option("--error-if-tagged-old-versions/--ignore-tagged-old-versions"),
    ] = None,
    delete_rate_limit: Annotated[
        Optional[int], typer.Option("--delete-rate-limit")
    ] = None,
    config_json: Annotated[
        Optional[str],
        typer.Option("--config-json", help="JSON object with optimization fields."),
    ] = None,
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
    o: OutputFormatOption = "table",
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            if config_json is not None:
                try:
                    config_data = _json.loads(config_json)
                except ValueError as exc:
                    raise typer.BadParameter(
                        "--config-json must be a JSON object",
                        param_hint="--config-json",
                    ) from exc
                if not isinstance(config_data, dict):
                    raise typer.BadParameter(
                        "--config-json must be a JSON object",
                        param_hint="--config-json",
                    )
            else:
                config_data = {}

            def add_config(key: str, value: Any) -> None:
                if value is not None:
                    config_data[key] = value

            add_config("compact_files", compact_files)
            add_config("optimize_indices", optimize_indices)
            add_config("cleanup_old_versions", cleanup_old_versions)
            config_data.setdefault("cleanup_old_versions", False)
            add_config("target_rows_per_fragment", target_rows_per_fragment)
            add_config("max_rows_per_group", max_rows_per_group)
            add_config("max_bytes_per_file", max_bytes_per_file)
            add_config("materialize_deletions", materialize_deletions)
            add_config(
                "materialize_deletions_threshold",
                materialize_deletions_threshold,
            )
            add_config("defer_index_remap", defer_index_remap)
            add_config("num_threads", num_threads)
            add_config("batch_size", batch_size)
            add_config("compaction_mode", compaction_mode)
            add_config("binary_copy_read_batch_bytes", binary_copy_read_batch_bytes)
            add_config("num_indices_to_merge", num_indices_to_merge)
            add_config("index_names", index_names)
            add_config("retrain", retrain)
            add_config("older_than_seconds", older_than_seconds)
            add_config("retain_versions", retain_versions)
            add_config("delete_unverified", delete_unverified)
            add_config("error_if_tagged_old_versions", error_if_tagged_old_versions)
            add_config("delete_rate_limit", delete_rate_limit)

            config = (
                DatasetOptimizeConfig.model_validate(config_data)
                if config_data
                else None
            )
            result = await client.datasets.optimize(
                table=table,
                namespace=_ns(namespace),
                branch=branch,
                config=config,
            )
            out = result.model_dump(mode="json")
            if o == "json":
                print(_json.dumps(out, indent=2))
            else:
                rows = []
                for group, values in out.items():
                    if isinstance(values, dict):
                        rows.extend(
                            {"group": group, "name": key, "value": value}
                            for key, value in values.items()
                        )
                    else:
                        rows.append(
                            {"group": "optimize", "name": group, "value": values}
                        )
                print_json_table(rows, "group", "name", "value")

    except (RoomException, typer.BadParameter, ValidationError) as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command("stats", help="Show statistics for a room dataset table.")
async def stats(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
    version: VersionOption = None,
    max_rows_per_group: Annotated[
        Optional[int], typer.Option("--max-rows-per-group")
    ] = None,
    o: OutputFormatOption = "table",
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            result = await client.datasets.stats(
                table=table,
                namespace=_ns(namespace),
                branch=branch,
                version=version,
                max_rows_per_group=max_rows_per_group,
            )
            out = result.model_dump(mode="json")
            if o == "json":
                print(_json.dumps(out, indent=2))
                return

            rows = []
            for group, values in out.items():
                rows.extend(
                    {"group": group, "name": key, "value": value}
                    for key, value in values.items()
                )
            print_json_table(rows, "group", "name", "value")

    except RoomException as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command(
    "versions",
    help="List versions for a room dataset table.",
    hidden=True,
)
@app.async_command("version", help="List versions for a room dataset table.")
async def list_versions(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
    pretty: Annotated[
        bool, typer.Option("--pretty/--no-pretty", help="Pretty-print JSON")
    ] = True,
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            versions = await client.datasets.list_versions(
                table=table,
                namespace=_ns(namespace),
                branch=branch,
            )
            out = [v.model_dump(mode="json") for v in versions]
            print(_json.dumps(out, indent=2 if pretty else None))

    except RoomException as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@branch_app.async_command("list", help="List dataset branches in a room namespace.")
async def list_branches(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    namespace: NamespaceOption = None,
    pretty: Annotated[
        bool, typer.Option("--pretty/--no-pretty", help="Pretty-print JSON")
    ] = True,
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            branches = await client.datasets.list_branches(namespace=_ns(namespace))
            out = [branch.model_dump(mode="json") for branch in branches]
            print(_json.dumps(out, indent=2 if pretty else None))

    except RoomException as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@branch_app.async_command("create", help="Create a dataset branch.")
async def create_branch(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    branch: Annotated[str, typer.Option(..., "--branch", help="New branch name")],
    from_branch: Annotated[
        Optional[str],
        typer.Option("--from-branch", help="Source branch to branch from"),
    ] = None,
    namespace: NamespaceOption = None,
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            await client.datasets.create_branch(
                branch=branch,
                from_branch=from_branch,
                namespace=_ns(namespace),
            )
            if from_branch is None:
                print(f"[bold green]Created branch[/bold green] {branch} from main")
            else:
                print(
                    f"[bold green]Created branch[/bold green] {branch} from {from_branch}"
                )

    except RoomException as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@branch_app.async_command("delete", help="Delete a dataset branch.")
async def delete_branch(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    branch: Annotated[str, typer.Option(..., "--branch", help="Branch name")],
    namespace: NamespaceOption = None,
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            await client.datasets.delete_branch(
                branch=branch,
                namespace=_ns(namespace),
            )
            print(f"[bold green]Deleted branch[/bold green] {branch}")

    except RoomException as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command(
    "restore", help="Restore a room dataset table to a specific version."
)
async def restore(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    version: Annotated[int, typer.Option(..., "--version", "-v", help="Table version")],
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            await client.datasets.restore(
                table=table,
                version=version,
                namespace=_ns(namespace),
                branch=branch,
            )
            print(f"[bold green]Restored[/bold green] {table} to version {version}")

    except RoomException as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command(
    "indexes",
    help="List indexes on a room dataset table.",
    hidden=True,
)
@app.async_command("index", help="List indexes on a room dataset table.")
async def list_indexes(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
    version: VersionOption = None,
    pretty: Annotated[
        bool, typer.Option("--pretty/--no-pretty", help="Pretty-print JSON")
    ] = True,
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            idxs = await client.datasets.list_indexes(
                table=table,
                namespace=_ns(namespace),
                branch=branch,
                version=version,
            )
            out = [i.model_dump(mode="json") for i in idxs]
            print(_json.dumps(out, indent=2 if pretty else None))

    except RoomException as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command("index-create", help="Create an index on a room dataset table.")
async def create_index(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    column: Annotated[
        Optional[List[str]],
        typer.Option(
            "--column",
            "-c",
            help="Column name. Repeat this option to pass multiple columns.",
        ),
    ] = None,
    index_type: Annotated[
        Optional[str],
        typer.Option(
            "--index-type",
            help=(
                "Lance index type, such as IVF_PQ, IVF_HNSW_PQ, IVF_HNSW_SQ, "
                "IVF_RQ, BTREE, BITMAP, LABEL_LIST, NGRAM, ZONEMAP, INVERTED, "
                "FTS, BLOOMFILTER, or RTREE."
            ),
        ),
    ] = None,
    name: Annotated[Optional[str], typer.Option("--name", help="Index name")] = None,
    metric: Annotated[
        Optional[str],
        typer.Option(
            "--metric", help="Vector distance metric, such as L2, cosine, or dot"
        ),
    ] = None,
    replace: Annotated[
        Optional[bool],
        typer.Option(
            "--replace/--no-replace",
            help="Replace existing index if it already exists",
        ),
    ] = None,
    train: Annotated[
        Optional[bool],
        typer.Option("--train/--no-train", help="Train the index on existing data"),
    ] = None,
    num_partitions: Annotated[Optional[int], typer.Option("--num-partitions")] = None,
    target_partition_size: Annotated[
        Optional[int], typer.Option("--target-partition-size")
    ] = None,
    num_sub_vectors: Annotated[Optional[int], typer.Option("--num-sub-vectors")] = None,
    num_bits: Annotated[Optional[int], typer.Option("--num-bits")] = None,
    accelerator: Annotated[Optional[str], typer.Option("--accelerator")] = None,
    index_cache_size: Annotated[
        Optional[int], typer.Option("--index-cache-size")
    ] = None,
    shuffle_partition_batches: Annotated[
        Optional[int], typer.Option("--shuffle-partition-batches")
    ] = None,
    shuffle_partition_concurrency: Annotated[
        Optional[int], typer.Option("--shuffle-partition-concurrency")
    ] = None,
    ivf_centroids_file: Annotated[
        Optional[str], typer.Option("--ivf-centroids-file")
    ] = None,
    precomputed_partition_dataset: Annotated[
        Optional[str], typer.Option("--precomputed-partition-dataset")
    ] = None,
    filter_nan: Annotated[
        Optional[bool],
        typer.Option(
            "--filter-nan/--no-filter-nan", help="Filter null or NaN vector values"
        ),
    ] = None,
    index_uuid: Annotated[Optional[str], typer.Option("--index-uuid")] = None,
    skip_transpose: Annotated[
        Optional[bool],
        typer.Option(
            "--skip-transpose/--no-skip-transpose",
            help="Skip vector index transposition",
        ),
    ] = None,
    index_file_version: Annotated[
        Optional[str], typer.Option("--index-file-version")
    ] = None,
    max_level: Annotated[Optional[int], typer.Option("--max-level")] = None,
    m: Annotated[Optional[int], typer.Option("--m")] = None,
    ef_construction: Annotated[Optional[int], typer.Option("--ef-construction")] = None,
    with_position: Annotated[
        Optional[bool],
        typer.Option(
            "--with-position/--no-with-position",
            help="Store token positions for text indexes",
        ),
    ] = None,
    memory_limit: Annotated[Optional[int], typer.Option("--memory-limit")] = None,
    num_workers: Annotated[Optional[int], typer.Option("--num-workers")] = None,
    skip_merge: Annotated[
        Optional[bool],
        typer.Option(
            "--skip-merge/--no-skip-merge", help="Skip text index partition merge"
        ),
    ] = None,
    base_tokenizer: Annotated[Optional[str], typer.Option("--base-tokenizer")] = None,
    language: Annotated[Optional[str], typer.Option("--language")] = None,
    max_token_length: Annotated[
        Optional[int], typer.Option("--max-token-length")
    ] = None,
    lower_case: Annotated[
        Optional[bool],
        typer.Option(
            "--lower-case/--no-lower-case", help="Lowercase text index tokens"
        ),
    ] = None,
    stem: Annotated[
        Optional[bool], typer.Option("--stem/--no-stem", help="Stem text index tokens")
    ] = None,
    remove_stop_words: Annotated[
        Optional[bool],
        typer.Option(
            "--remove-stop-words/--keep-stop-words", help="Remove text index stop words"
        ),
    ] = None,
    custom_stop_words: Annotated[
        Optional[List[str]],
        typer.Option(
            "--custom-stop-word",
            help="Custom stop word. Repeat to pass multiple words.",
        ),
    ] = None,
    ascii_folding: Annotated[
        Optional[bool],
        typer.Option(
            "--ascii-folding/--no-ascii-folding", help="Fold text index tokens to ASCII"
        ),
    ] = None,
    config_json: Annotated[
        Optional[str],
        typer.Option(
            "--config-json", help="JSON object with Lance index configuration fields."
        ),
    ] = None,
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            if config_json is not None:
                try:
                    config_data = _json.loads(config_json)
                except ValueError as exc:
                    raise typer.BadParameter(
                        "--config-json must be a JSON object",
                        param_hint="--config-json",
                    ) from exc
                if not isinstance(config_data, dict):
                    raise typer.BadParameter(
                        "--config-json must be a JSON object",
                        param_hint="--config-json",
                    )
            else:
                if not column:
                    raise typer.BadParameter(
                        "--column is required when --config-json is not provided",
                        param_hint="--column",
                    )
                if index_type is None:
                    raise typer.BadParameter(
                        "--index-type is required when --config-json is not provided",
                        param_hint="--index-type",
                    )
                selected_column: str | list[str] = (
                    column[0] if len(column) == 1 else column
                )
                config_data = {"column": selected_column, "index_type": index_type}

            def add_config(key: str, value: Any) -> None:
                if value is not None:
                    config_data[key] = value

            add_config("name", name)
            add_config("metric", metric)
            add_config("replace", replace)
            add_config("train", train)
            add_config("num_partitions", num_partitions)
            add_config("target_partition_size", target_partition_size)
            add_config("num_sub_vectors", num_sub_vectors)
            add_config("num_bits", num_bits)
            add_config("accelerator", accelerator)
            add_config("index_cache_size", index_cache_size)
            add_config("shuffle_partition_batches", shuffle_partition_batches)
            add_config("shuffle_partition_concurrency", shuffle_partition_concurrency)
            add_config("ivf_centroids_file", ivf_centroids_file)
            add_config("precomputed_partition_dataset", precomputed_partition_dataset)
            add_config("filter_nan", filter_nan)
            add_config("index_uuid", index_uuid)
            add_config("skip_transpose", skip_transpose)
            add_config("index_file_version", index_file_version)
            add_config("max_level", max_level)
            add_config("m", m)
            add_config("ef_construction", ef_construction)
            add_config("with_position", with_position)
            add_config("memory_limit", memory_limit)
            add_config("num_workers", num_workers)
            add_config("skip_merge", skip_merge)
            add_config("base_tokenizer", base_tokenizer)
            add_config("language", language)
            add_config("max_token_length", max_token_length)
            add_config("lower_case", lower_case)
            add_config("stem", stem)
            add_config("remove_stop_words", remove_stop_words)
            add_config("custom_stop_words", custom_stop_words)
            add_config("ascii_folding", ascii_folding)

            config = DatasetIndexConfig.model_validate(config_data)
            await client.datasets.create_index(
                table=table,
                config=config,
                namespace=_ns(namespace),
                branch=branch,
            )

            print(
                f"[bold green]Created[/bold green] {config.index_type} index on {table}"
            )

    except (RoomException, typer.BadParameter) as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()


@app.async_command("index-drop", help="Drop an index from a room dataset table.")
async def drop_index(
    *,
    project_id: ProjectIdOption,
    room: RoomOption,
    table: Annotated[str, typer.Option(..., "--table", "-t", help="Table name")],
    name: Annotated[str, typer.Option(..., "--name", help="Index name")],
    namespace: NamespaceOption = None,
    branch: BranchOption = None,
):
    account_client = await get_client()
    try:
        project_id = await resolve_project_id(project_id=project_id)
        room_name = resolve_room(room)
        connection = await account_client.connect_room(
            project_id=project_id, room=room_name
        )

        async with RoomClient(
            protocol_factory=WebSocketClientProtocol(
                url=websocket_room_url(room_name=room_name),
                token=connection.jwt,
            ).create_factory()
        ) as client:
            await client.datasets.drop_index(
                table=table,
                name=name,
                namespace=_ns(namespace),
                branch=branch,
            )
            print(f"[bold green]Dropped index[/bold green] {name} on {table}")

    except RoomException as e:
        print(e)
        raise typer.Exit(1)
    finally:
        await account_client.close()
